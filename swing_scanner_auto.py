"""
Automated Swing Trading Scanner
================================
Features:
  - Nifty 100 stocks scanned every 15 minutes during market hours
  - Indicators : RSI, MACD, EMA, Bollinger Bands, ADX
  - Filters    : Volume spike, Nifty 50 trend, News sentiment
  - S/R        : Swing highs/lows, 52-week high/low, round numbers
  - Candles    : Hammer, Engulfing, Morning/Evening Star, Shooting Star
  - Stop Loss  : ATR-based dynamic
  - Scoring    : Out of 13 points, High confidence = 10+
  - Stock trend: Per-stock daily EMA filter (stock must align with signal)
  - Output     : Console table, no duplicate signals per day

Requirements:
    pip install yfinance pandas pandas-ta requests rich feedparser pytz openpyxl

Run:
    python swing_scanner_auto.py
"""

import warnings
warnings.filterwarnings("ignore")

import yfinance as yf
import pandas as pd
#import pandas_ta as ta
import numpy as np
import feedparser
import time
import pytz
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime, date
from rich.console import Console
from rich.table import Table
from rich.text import Text
from rich import box

# Excel log file — saved in same folder as this script
EXCEL_LOG_FILE = "swing_signals_log.xlsx"

SCAN_INTERVAL_SECONDS = 15 * 60
IST                   = pytz.timezone("Asia/Kolkata")
MARKET_OPEN           = (9, 15)
MARKET_CLOSE          = (15, 30)

# Indicator settings
RSI_PERIOD     = 14
RSI_OVERSOLD   = 40
BB_PERIOD      = 20
BB_STD         = 2
EMA_SHORT      = 20
EMA_LONG       = 50
ADX_PERIOD     = 14
ADX_THRESHOLD  = 25
ATR_PERIOD     = 14
VOLUME_MULT    = 1.7
SR_ZONE_PCT    = 0.015
RISK_REWARD    = 2.0
MIN_SCORE      = 9    # Raised from 8 → only best setups
ATR_MULT       = 2.0    # Raised from 1.5 → wider stop, less noise

console = Console()

# ─────────────────────────────────────────────
# NIFTY 100 SYMBOLS
# ─────────────────────────────────────────────
NIFTY100_SYMBOLS = [
    "RELIANCE", "TCS", "HDFCBANK", "INFY", "ICICIBANK",
    "HINDUNILVR", "ITC", "SBIN", "BHARTIARTL", "KOTAKBANK",
    "LT", "AXISBANK", "ASIANPAINT", "MARUTI", "WIPRO",
    "HCLTECH", "ULTRACEMCO", "BAJFINANCE", "NESTLEIND", "TITAN",
    "TECHM", "SUNPHARMA", "ONGC", "NTPC", "POWERGRID",
    "M&M", "BAJAJFINSV", "TATAMOTORS", "TATASTEEL", "ADANIENT",
    "JSWSTEEL", "HINDALCO", "COALINDIA", "DRREDDY", "CIPLA",
    "DIVISLAB", "GRASIM", "HEROMOTOCO", "EICHERMOT", "BPCL",
    "BRITANNIA", "INDUSINDBK", "APOLLOHOSP", "TATACONSUM", "DABUR",
    "PIDILITIND", "SIEMENS", "HAVELLS", "BERGEPAINT", "GODREJCP",
    "MARICO", "MUTHOOTFIN", "BANDHANBNK", "PNB", "CANBK",
    "BANKBARODA", "LUPIN", "BIOCON", "TORNTPHARM", "IPCALAB",
    "GLAND", "ALKEM", "AUROPHARMA", "ZYDUSLIFE", "ABBOTINDIA",
    "SANOFI", "PFIZER", "GLAXO", "CHOLAFIN", "BAJAJ-AUTO",
    "TVSMOTOR", "MOTHERSON", "BOSCHLTD", "BALKRISIND", "MRF",
    "CUMMINSIND", "ABB", "BHEL", "CONCOR", "ADANIPORTS",
    "GMRINFRA", "IRCTC", "DMART", "NYKAA", "ZOMATO",
    "PAYTM", "POLICYBZR", "NAVINFLUOR", "SRF", "AAVAS",
    "CROMPTON", "VOLTAS", "WHIRLPOOL", "BLUESTAR", "VGUARD",
    "HFCL", "RVNL", "IRFC", "RAILTEL", "HUDCO"
]


# ─────────────────────────────────────────────
# NEWS SENTIMENT
# ─────────────────────────────────────────────
def get_news_sentiment(symbol: str) -> tuple:
    positive_words = [
        "surge", "rally", "gain", "profit", "growth", "beat", "upgrade",
        "outperform", "buy", "strong", "record", "boost", "expansion",
        "deal", "order", "positive", "rise", "bullish", "momentum",
        "breakout", "revenue", "acquisition"
    ]
    negative_words = [
        "fall", "drop", "loss", "decline", "downgrade", "miss", "weak",
        "sell", "cut", "reduce", "concern", "risk", "fraud", "penalty",
        "fine", "negative", "bearish", "crash", "debt", "warning",
        "lawsuit", "probe", "slump", "plunge"
    ]
    try:
        url  = f"https://news.google.com/rss/search?q={symbol}+NSE+stock&hl=en-IN&gl=IN&ceid=IN:en"
        feed = feedparser.parse(url)
        entries = feed.entries[:10]
        if not entries:
            return 0.0, "Neutral"
        score = 0
        for entry in entries:
            text = (entry.get("title", "") + " " + entry.get("summary", "")).lower()
            for w in positive_words:
                if w in text: score += 1
            for w in negative_words:
                if w in text: score -= 1
        normalized = score / max(len(entries), 1)
        if normalized > 0.3:    label = "Positive"
        elif normalized < -0.3: label = "Negative"
        else:                   label = "Neutral"
        return round(normalized, 2), label
    except Exception:
        return 0.0, "Neutral"


# ─────────────────────────────────────────────
# NIFTY 50 TREND
# ─────────────────────────────────────────────
def get_nifty_trend() -> str:
    try:
        df = yf.download("^NSEI", period="3mo", interval="1d",
                         progress=False, auto_adjust=True)
        df.columns = [c[0] if isinstance(c, tuple) else c for c in df.columns]
        df.dropna(inplace=True)
        df.ta.ema(length=20, append=True)
        df.ta.ema(length=50, append=True)
        last  = df.iloc[-1]
        ema20 = float(last.get("EMA_20", 0))
        ema50 = float(last.get("EMA_50", 0))
        close = float(last["Close"])
        if close > ema20 > ema50:   return "bullish"
        elif close < ema20 < ema50: return "bearish"
        else:                       return "neutral"
    except Exception:
        return "neutral"


# ─────────────────────────────────────────────
# IMPROVEMENT 2 — PER-STOCK DAILY TREND FILTER
# ─────────────────────────────────────────────
def get_stock_trend(symbol: str) -> str:
    """
    Checks the stock's own daily trend using EMA20 vs EMA50.
    A BUY signal on a stock in a daily downtrend is risky even
    if Nifty is bullish. This filter catches that.
    Returns 'bullish', 'bearish', or 'neutral'.
    """
    try:
        ticker = symbol + ".NS"
        df = yf.download(ticker, period="3mo", interval="1d",
                         progress=False, auto_adjust=True)
        df.columns = [c[0] if isinstance(c, tuple) else c for c in df.columns]
        df.dropna(inplace=True)
        df.ta.ema(length=20, append=True)
        df.ta.ema(length=50, append=True)
        last  = df.iloc[-1]
        ema20 = float(last.get("EMA_20", 0))
        ema50 = float(last.get("EMA_50", 0))
        close = float(last["Close"])
        if close > ema20 > ema50:   return "bullish"
        elif close < ema20 < ema50: return "bearish"
        else:                       return "neutral"
    except Exception:
        return "neutral"


# ─────────────────────────────────────────────
# SUPPORT / RESISTANCE
# ─────────────────────────────────────────────
def find_sr_levels(df: pd.DataFrame, close: float) -> tuple:
    levels = []

    # Swing highs/lows
    highs = df["High"].rolling(5, center=True).max()
    lows  = df["Low"].rolling(5, center=True).min()
    levels.extend(df["High"][df["High"] == highs].tail(50).tolist())
    levels.extend(df["Low"][df["Low"] == lows].tail(50).tolist())

    # 52-week high/low
    levels.append(float(df["High"].tail(252 * 26).max()))
    levels.append(float(df["Low"].tail(252 * 26).min()))

    # Round numbers
    for base in [50, 100]:
        for offset in [-1, 0, 1]:
            levels.append((round(close / base) + offset) * base)

    zone = close * SR_ZONE_PCT
    near_support    = any(abs(close - l) <= zone and l <= close for l in levels if l > 0)
    near_resistance = any(abs(close - l) <= zone and l >= close for l in levels if l > 0)
    return near_support, near_resistance


# ─────────────────────────────────────────────
# CANDLESTICK PATTERNS
# ─────────────────────────────────────────────
def detect_candle_pattern(df: pd.DataFrame) -> tuple:
    if len(df) < 3:
        return False, False, ""

    c1 = df.iloc[-3]
    c2 = df.iloc[-2]
    c3 = df.iloc[-1]

    o3, h3, l3, c3c = float(c3["Open"]), float(c3["High"]), float(c3["Low"]), float(c3["Close"])
    o2, h2, l2, c2c = float(c2["Open"]), float(c2["High"]), float(c2["Low"]), float(c2["Close"])
    o1, h1, l1, c1c = float(c1["Open"]), float(c1["High"]), float(c1["Low"]), float(c1["Close"])

    body3         = abs(c3c - o3)
    range3        = h3 - l3 if h3 != l3 else 0.001
    upper_wick3   = h3 - max(c3c, o3)
    lower_wick3   = min(c3c, o3) - l3

    # Bullish
    hammer        = lower_wick3 >= 2 * body3 and upper_wick3 <= 0.1 * range3 and body3 > 0
    bull_engulf   = c2c < o2 and c3c > o3 and o3 <= c2c and c3c >= o2
    morning_star  = (c1c < o1 and
                     abs(c2c - o2) < 0.3 * (h2 - l2 if h2 != l2 else 0.001) and
                     c3c > o3 and c3c > (o1 + c1c) / 2)

    # Bearish
    shooting_star = upper_wick3 >= 2 * body3 and lower_wick3 <= 0.1 * range3 and body3 > 0
    bear_engulf   = c2c > o2 and c3c < o3 and o3 >= c2c and c3c <= o2
    evening_star  = (c1c > o1 and
                     abs(c2c - o2) < 0.3 * (h2 - l2 if h2 != l2 else 0.001) and
                     c3c < o3 and c3c < (o1 + c1c) / 2)

    bullish = hammer or bull_engulf or morning_star
    bearish = shooting_star or bear_engulf or evening_star

    if hammer:          name = "Hammer"
    elif bull_engulf:   name = "Bullish Engulfing"
    elif morning_star:  name = "Morning Star"
    elif shooting_star: name = "Shooting Star"
    elif bear_engulf:   name = "Bearish Engulfing"
    elif evening_star:  name = "Evening Star"
    else:               name = ""

    return bullish, bearish, name


# ─────────────────────────────────────────────
# CORE STOCK ANALYSIS
# ─────────────────────────────────────────────
def analyze_stock(symbol: str, nifty_trend: str) -> dict | None:
    ticker = symbol + ".NS"
    try:
        df = yf.download(ticker, period="60d", interval="15m",
                         progress=False, auto_adjust=True)
        if df is None or len(df) < EMA_LONG + 20:
            return None

        df.columns = [c[0] if isinstance(c, tuple) else c for c in df.columns]
        df.dropna(inplace=True)

        df.ta.rsi(length=RSI_PERIOD, append=True)
        df.ta.macd(fast=12, slow=26, signal=9, append=True)
        df.ta.bbands(length=BB_PERIOD, std=BB_STD, append=True)
        df.ta.ema(length=EMA_SHORT, append=True)
        df.ta.ema(length=EMA_LONG, append=True)
        df.ta.adx(length=ADX_PERIOD, append=True)
        df.ta.atr(length=ATR_PERIOD, append=True)

        df.dropna(inplace=True)
        if len(df) < 5:
            return None

        last = df.iloc[-1]
        prev = df.iloc[-2]

        close     = float(last["Close"])
        volume    = float(last["Volume"])
        rsi       = float(last.get(f"RSI_{RSI_PERIOD}", 50))
        macd_hist = float(last.get("MACDh_12_26_9", 0))
        prev_macd = float(prev.get("MACDh_12_26_9", 0))
        bb_lower  = float(last.get(f"BBL_{BB_PERIOD}_{float(BB_STD)}.0", close))
        bb_upper  = float(last.get(f"BBU_{BB_PERIOD}_{float(BB_STD)}.0", close))
        ema_short = float(last.get(f"EMA_{EMA_SHORT}", close))
        ema_long  = float(last.get(f"EMA_{EMA_LONG}", close))
        adx       = float(last.get(f"ADX_{ADX_PERIOD}", 0))
        atr       = float(last.get(f"ATRr_{ATR_PERIOD}", close * 0.01))

        # Volume filter — must be > 1.5x average
        avg_vol = df["Volume"].tail(20 * 26).mean()
        if volume <= avg_vol * VOLUME_MULT:
            return None

        buy_score  = 0
        sell_score = 0
        reasons    = []

        # 1. RSI (+2)
        if rsi < RSI_OVERSOLD:
            buy_score += 2; reasons.append(f"RSI({rsi:.0f}) oversold")
        elif rsi > (100 - RSI_OVERSOLD):
            sell_score += 2; reasons.append(f"RSI({rsi:.0f}) overbought")

        # 2. MACD (+2)
        if macd_hist > 0 and prev_macd <= 0:
            buy_score += 2; reasons.append("MACD ↑ cross")
        elif macd_hist < 0 and prev_macd >= 0:
            sell_score += 2; reasons.append("MACD ↓ cross")
        elif macd_hist > 0:
            buy_score += 1
        elif macd_hist < 0:
            sell_score += 1

        # 3. Bollinger Bands (+2)
        if close <= bb_lower:
            buy_score += 2; reasons.append("At BB lower")
        elif close >= bb_upper:
            sell_score += 2; reasons.append("At BB upper")

        # 4. EMA trend (+1)
        if ema_short > ema_long:
            buy_score += 1; reasons.append("EMA bullish")
        else:
            sell_score += 1; reasons.append("EMA bearish")

        # 5. ADX (+2)
        if adx > ADX_THRESHOLD:
            if ema_short > ema_long:
                buy_score += 2
            else:
                sell_score += 2
            reasons.append(f"ADX({adx:.0f}) strong")

        # 6. Support / Resistance (+2)
        near_sup, near_res = find_sr_levels(df, close)
        if near_sup:
            buy_score += 2; reasons.append("Near support")
        if near_res:
            sell_score += 2; reasons.append("Near resistance")

        # 7. Candlestick pattern (+2)
        bull_c, bear_c, c_name = detect_candle_pattern(df)
        if bull_c:
            buy_score += 2; reasons.append(c_name)
        elif bear_c:
            sell_score += 2; reasons.append(c_name)

        # Determine signal
        if buy_score >= MIN_SCORE:
            signal, score = "BUY", buy_score
        elif sell_score >= MIN_SCORE:
            signal, score = "SELL", sell_score
        else:
            return None

        # Nifty trend filter
        if signal == "BUY"  and nifty_trend == "bearish": return None
        if signal == "SELL" and nifty_trend == "bullish": return None

        # Improvement 2 — Per-stock daily trend filter
        stock_trend = get_stock_trend(symbol)
        if signal == "BUY"  and stock_trend == "bearish": return None
        if signal == "SELL" and stock_trend == "bullish": return None

        # News sentiment filter
        _, sentiment_label = get_news_sentiment(symbol)
        if signal == "BUY"  and sentiment_label == "Negative": return None
        if signal == "SELL" and sentiment_label == "Positive": return None

        # Improvement 3 — ATR multiplier raised to 2.0 (wider stop, less noise)
        if signal == "BUY":
            stop_loss = round(close - ATR_MULT * atr, 2)
            target    = round(close + (close - stop_loss) * RISK_REWARD, 2)
        else:
            stop_loss = round(close + ATR_MULT * atr, 2)
            target    = round(close - (stop_loss - close) * RISK_REWARD, 2)

        return {
            "symbol":    symbol,
            "signal":    signal,
            "entry":     round(close, 2),
            "target":    target,
            "stop_loss": stop_loss,
            "score":     score,
            "rsi":       round(rsi, 1),
            "adx":       round(adx, 1),
            "sentiment": sentiment_label,
            "candle":    c_name or "—",
            "reasons":   " | ".join(reasons),
        }

    except Exception:
        return None


# ─────────────────────────────────────────────
# CONSOLE TABLE
# ─────────────────────────────────────────────
def print_results(results: list, scan_time: str):
    if not results:
        console.print(f"[dim]{scan_time} — No high-confidence signals this scan.[/dim]")
        return

    table = Table(
        title=f"Swing Trade Signals  |  {scan_time}",
        box=box.ROUNDED, show_lines=True,
        header_style="bold white on dark_blue",
    )
    table.add_column("Symbol",    style="bold", width=14)
    table.add_column("Signal",    justify="center", width=8)
    table.add_column("Entry ₹",  justify="right",  width=10)
    table.add_column("Target ₹", justify="right",  width=10)
    table.add_column("SL ₹",     justify="right",  width=10)
    table.add_column("Score",     justify="center", width=7)
    table.add_column("RSI",       justify="right",  width=6)
    table.add_column("ADX",       justify="right",  width=6)
    table.add_column("Candle",    width=18)
    table.add_column("News",      justify="center", width=10)
    table.add_column("Reasons",   width=38)

    for r in results:
        sig_c  = "bold green" if r["signal"] == "BUY" else "bold red"
        tgt_c  = "green"      if r["signal"] == "BUY" else "red"
        sl_c   = "red"        if r["signal"] == "BUY" else "green"
        news_c = {"Positive": "green", "Negative": "red"}.get(r["sentiment"], "yellow")

        table.add_row(
            r["symbol"],
            Text(r["signal"], style=sig_c),
            f"{r['entry']:,.2f}",
            Text(f"{r['target']:,.2f}", style=tgt_c),
            Text(f"{r['stop_loss']:,.2f}", style=sl_c),
            f"{r['score']}/13",
            str(r["rsi"]),
            str(r["adx"]),
            r["candle"],
            Text(r["sentiment"], style=news_c),
            Text(r["reasons"], style="dim"),
        )
    console.print(table)


# ─────────────────────────────────────────────
# EXCEL SIGNAL LOGGER
# ─────────────────────────────────────────────
def log_to_excel(signal_data: dict, scan_time: str):
    """
    Appends a new signal row to the Excel log file.
    Creates the file with formatted headers if it doesn't exist yet.
    Columns: Date | Time | Stock | Signal | Entry | Target | Stop Loss | Score | R/R | Candle | Reasons
    """
    file_path = Path(EXCEL_LOG_FILE)

    # Colour palette
    GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
    RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    BODY_FONT   = Font(size=10)
    CENTER      = Alignment(horizontal="center", vertical="center")
    LEFT        = Alignment(horizontal="left",   vertical="center")
    THIN        = Side(style="thin", color="BFBFBF")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    HEADERS    = ["Date", "Time", "Stock", "Signal",
                  "Entry ₹", "Target ₹", "Stop Loss ₹",
                  "Score", "R/R", "Candle Pattern", "Reasons"]
    COL_WIDTHS = [13, 8, 14, 8, 12, 12, 14, 8, 6, 18, 45]

    # Load or create workbook
    if file_path.exists():
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Signals"

        # Write header row
        for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            cell           = ws.cell(row=1, column=col_idx, value=header)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = CENTER
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

    # Parse date and time from scan_time string
    try:
        dt       = datetime.strptime(scan_time[:16], "%d %b %Y %H:%M")
        date_str = dt.strftime("%d-%m-%Y")
        time_str = dt.strftime("%H:%M")
    except Exception:
        date_str = scan_time[:10]
        time_str = scan_time[11:16]

    # Compute actual R/R for this trade
    risk   = abs(signal_data["entry"] - signal_data["stop_loss"])
    reward = abs(signal_data["target"] - signal_data["entry"])
    rr     = f"1 : {round(reward / risk, 1)}" if risk > 0 else "N/A"

    row_values = [
        date_str,
        time_str,
        signal_data["symbol"],
        signal_data["signal"],
        signal_data["entry"],
        signal_data["target"],
        signal_data["stop_loss"],
        f"{signal_data['score']}/13",
        rr,
        signal_data.get("candle", "—"),
        signal_data.get("reasons", ""),
    ]

    # Write data row
    next_row = ws.max_row + 1
    row_fill = GREEN_FILL if signal_data["signal"] == "BUY" else RED_FILL

    for col_idx, value in enumerate(row_values, 1):
        cell           = ws.cell(row=next_row, column=col_idx, value=value)
        cell.font      = BODY_FONT
        cell.border    = BORDER
        cell.fill      = row_fill

        if col_idx in (1, 2, 4, 8, 9, 10):      # center: Date, Time, Signal, Score, R/R, Candle
            cell.alignment = CENTER
        elif col_idx in (5, 6, 7):               # right: price columns
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
        else:
            cell.alignment = LEFT

    ws.row_dimensions[next_row].height = 16

    try:
        wb.save(file_path)
    except PermissionError:
        console.print(
            f"[yellow]  ⚠ Could not save Excel — close {EXCEL_LOG_FILE} and it will retry next scan.[/yellow]"
        )


# ─────────────────────────────────────────────
# MARKET HOURS
# ─────────────────────────────────────────────
def is_market_open() -> bool:
    now = datetime.now(IST)
    if now.weekday() >= 5:
        return False
    t = (now.hour, now.minute)
    return MARKET_OPEN <= t <= MARKET_CLOSE


# ─────────────────────────────────────────────
# MAIN LOOP
# ─────────────────────────────────────────────
def main():
    console.rule("[bold cyan]📈 Swing Trading Scanner — Auto Mode[/bold cyan]")
    console.print(
        "[dim]Nifty 100 | 15-min candles | RSI · MACD · EMA · BB · ADX · S/R · Candles · Volume · News[/dim]\n"
    )

    alerted_today:  dict = {}
    last_alert_date = date.today()

    while True:
        today = date.today()
        if today != last_alert_date:
            alerted_today   = {}
            last_alert_date = today

        if not is_market_open():
            now = datetime.now(IST).strftime("%d %b %Y %H:%M IST")
            console.print(f"[dim]{now} — Market closed. Checking again in 5 mins...[/dim]", end="\r")
            time.sleep(300)
            continue

        scan_time   = datetime.now(IST).strftime("%d %b %Y %H:%M IST")
        console.print(f"\n[bold]🔍 Scan started at {scan_time}[/bold]")

        nifty_trend = get_nifty_trend()
        console.print(f"[dim]Nifty 50 trend: {nifty_trend.upper()}[/dim]")

        results = []
        total   = len(NIFTY100_SYMBOLS)

        for i, symbol in enumerate(NIFTY100_SYMBOLS, 1):
            console.print(f"[dim]  Analyzing [{i}/{total}] {symbol}...[/dim]", end="\r")
            data = analyze_stock(symbol, nifty_trend)

            if data:
                key = f"{symbol}_{data['signal']}"
                if key not in alerted_today:
                    alerted_today[key] = scan_time
                    results.append(data)
                    log_to_excel(data, scan_time)
                    console.print(
                        f"[green]  🚨 {symbol} {data['signal']} | Score:{data['score']}/13 | ✅ Logged to Excel[/green]"
                    )

        console.print(" " * 70, end="\r")
        print_results(results, scan_time)
        console.print(f"[dim]Next scan in {SCAN_INTERVAL_SECONDS // 60} minutes...[/dim]")
        time.sleep(SCAN_INTERVAL_SECONDS)


if __name__ == "__main__":
    main()
